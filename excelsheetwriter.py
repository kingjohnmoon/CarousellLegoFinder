import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from carousellscraper import CarousellScraper
from bricklinkcomparer import BrickLinkComparer

class ExcelSheetWriter:
    def __init__(self, products, filename="lego_comparison.xlsx"):
        self.products = products
        self.filename = filename

    def write_to_excel(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lego Comparison"
        # Write header
        headers = ["Title", "Carousell Price", "Set Code", "BrickLink Price"]
        ws.append(headers)
        # Write product rows
        for product in self.products:
            ws.append([
                product.get("title", ""),
                product.get("price", 0.0),
                product.get("code", ""),
                product.get("bricklink_price", 0.0)
            ])
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2 

        # Apply conditional formatting to Carousell Price column (column B)
        for row in range(2, ws.max_row + 1):
            carousell_price = ws[f'B{row}'].value
            bricklink_price = ws[f'D{row}'].value
            if carousell_price is None or bricklink_price is None:
                continue
            try:
                diff = bricklink_price - carousell_price
                if diff < 0:
                    ws[f'B{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
                elif diff < 50:
                    ws[f'B{row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
                else:
                    ws[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
            except Exception:
                pass
        wb.save(self.filename)
        print(f"Excel file saved as {self.filename}")


if __name__ == "__main__":
    # Scrape products from Carousell
    scraper = CarousellScraper(1)  # Set the number of 'Brand new' listings to scrape here
    products = scraper.run("lego")
    
    # Compare with BrickLink prices
    comparer = BrickLinkComparer(products)
    cleaned_products = comparer.clean_products()
    price_info = comparer.get_price_info()
    filtered_products = comparer.filter_products_with_code_and_price()
    
    # Write to Excel
    writer = ExcelSheetWriter(filtered_products)
    writer.write_to_excel()